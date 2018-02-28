import json

import os

import pytest
from openpyxl import load_workbook

from entity.user import User
from fixture.application import Application

fixture = None
target = None

ROOT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)))

def load_config(file):
    global target
    if target is None:
        config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), file)
        with open(config_file) as file:
            target = json.load(file)
    return target

# start fixture with assertion on fixture valid
@pytest.fixture
# @pytest.fixture(scope="session") # use one browser for all test, BUT IT NEEDS TO ADD LOG OUT TO TEST
def app(request):
    global fixture

    browser = request.config.getoption("--browser")
    users = load_from_excel(request.config.getoption("--target"))

    if fixture is None or not fixture.is_valid():
        fixture = Application(browser=browser)

    # fixture.loginPage.login(username=users[0].username, password=users[0].password)
    return fixture

# log-out fixture. runs one time after all test
@pytest.fixture(scope="session", autouse=True)
def stop(request):
    def log_out():
        fixture.destroy()
    request.addfinalizer(log_out)
    return fixture

def pytest_addoption(parser):
    parser.addoption("--browser", action="store", default="chrome")
    parser.addoption("--target", action="store", default="credentials")

def pytest_generate_tests(metafunc):  # add dynamic data binding from file data/users.py
    for fixt in metafunc.fixturenames:
        if fixt.startswith("excel_"):
            testData = load_from_excel(fixt[6:])
            metafunc.parametrize(fixt, testData, ids=[str(x) for x in testData])


def load_from_excel(file):
    excel_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "%s.xlsx" % file)
    wb = load_workbook(filename=excel_file, read_only=True)
    ws = wb["Sheet1"]
    users = []
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=2, max_col=2):
        users.append(User(username=row[0].value, password=row[1].value))
    return users